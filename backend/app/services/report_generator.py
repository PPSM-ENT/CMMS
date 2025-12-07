"""
Report generation service for CMMS.
Generates PDF and Excel reports for work orders, assets, PM, labor, and inventory.
"""
import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """Generates PDF and Excel reports."""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom paragraph styles."""
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20,
        ))
        self.styles.add(ParagraphStyle(
            name='ReportSubtitle',
            parent=self.styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.gray,
            spaceAfter=30,
        ))
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceBefore=15,
            spaceAfter=10,
        ))
        self.styles.add(ParagraphStyle(
            name='MetricLabel',
            parent=self.styles['Normal'],
            fontSize=9,
            textColor=colors.gray,
        ))
        self.styles.add(ParagraphStyle(
            name='MetricValue',
            parent=self.styles['Normal'],
            fontSize=14,
            fontName='Helvetica-Bold',
        ))

    def _format_value(self, value: Any) -> str:
        """Format value for display."""
        if value is None:
            return '-'
        if isinstance(value, (float, Decimal)):
            if abs(value) >= 1000:
                return f'{value:,.2f}'
            return f'{value:.2f}'
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        return str(value)

    def _create_table_style(self) -> TableStyle:
        """Create standard table style."""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ])

    def generate_pdf(
        self,
        title: str,
        subtitle: str,
        summary_metrics: Optional[List[Dict[str, Any]]] = None,
        sections: Optional[List[Dict[str, Any]]] = None,
        landscape_mode: bool = False,
    ) -> bytes:
        """
        Generate a PDF report.

        Args:
            title: Report title
            subtitle: Report subtitle (date range, filters, etc.)
            summary_metrics: List of metric dicts with 'label' and 'value'
            sections: List of section dicts with 'title', 'headers', 'rows'
        """
        buffer = io.BytesIO()
        page_size = landscape(letter) if landscape_mode else letter
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
        )

        elements = []

        # Title
        elements.append(Paragraph(title, self.styles['ReportTitle']))
        elements.append(Paragraph(subtitle, self.styles['ReportSubtitle']))

        # Summary metrics
        if summary_metrics:
            metric_data = []
            labels = []
            values = []
            for metric in summary_metrics:
                labels.append(Paragraph(metric['label'], self.styles['MetricLabel']))
                values.append(Paragraph(self._format_value(metric['value']), self.styles['MetricValue']))

            if labels:
                metric_data = [labels, values]
                col_width = (page_size[0] - inch) / len(labels)
                metric_table = Table(metric_data, colWidths=[col_width] * len(labels))
                metric_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                    ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9fafb')),
                    ('TOPPADDING', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ]))
                elements.append(metric_table)
                elements.append(Spacer(1, 20))

        # Sections with tables
        if sections:
            for section in sections:
                if section.get('title'):
                    elements.append(Paragraph(section['title'], self.styles['SectionHeader']))

                if section.get('headers') and section.get('rows'):
                    # Prepare table data
                    headers = section['headers']
                    table_data = [headers]

                    for row in section['rows']:
                        formatted_row = [self._format_value(cell) for cell in row]
                        table_data.append(formatted_row)

                    # Calculate column widths
                    available_width = page_size[0] - inch
                    col_widths = section.get('col_widths')
                    if not col_widths:
                        col_widths = [available_width / len(headers)] * len(headers)

                    table = Table(table_data, colWidths=col_widths, repeatRows=1)
                    table.setStyle(self._create_table_style())

                    # Apply right alignment for numeric columns
                    numeric_cols = section.get('numeric_cols', [])
                    for col_idx in numeric_cols:
                        table.setStyle(TableStyle([
                            ('ALIGN', (col_idx, 1), (col_idx, -1), 'RIGHT'),
                        ]))

                    elements.append(table)
                    elements.append(Spacer(1, 15))

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_excel(
        self,
        title: str,
        subtitle: str,
        summary_metrics: Optional[List[Dict[str, Any]]] = None,
        sections: Optional[List[Dict[str, Any]]] = None,
    ) -> bytes:
        """
        Generate an Excel report.

        Args:
            title: Report title
            subtitle: Report subtitle (date range, filters, etc.)
            summary_metrics: List of metric dicts with 'label' and 'value'
            sections: List of section dicts with 'title', 'headers', 'rows'
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Styles
        title_font = Font(name='Arial', size=16, bold=True)
        subtitle_font = Font(name='Arial', size=10, color='666666')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        metric_label_font = Font(name='Arial', size=9, color='666666')
        metric_value_font = Font(name='Arial', size=12, bold=True)
        cell_font = Font(name='Arial', size=9)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        row_num = 1

        # Title
        ws.cell(row=row_num, column=1, value=title).font = title_font
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        row_num += 1

        # Subtitle
        ws.cell(row=row_num, column=1, value=subtitle).font = subtitle_font
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        row_num += 2

        # Summary metrics
        if summary_metrics:
            for idx, metric in enumerate(summary_metrics):
                col = idx + 1
                ws.cell(row=row_num, column=col, value=metric['label']).font = metric_label_font
                ws.cell(row=row_num + 1, column=col, value=metric['value']).font = metric_value_font
            row_num += 3

        # Sections with tables
        if sections:
            for section in sections:
                if section.get('title'):
                    ws.cell(row=row_num, column=1, value=section['title']).font = Font(
                        name='Arial', size=11, bold=True
                    )
                    row_num += 1

                if section.get('headers') and section.get('rows'):
                    # Headers
                    for col_idx, header in enumerate(section['headers'], 1):
                        cell = ws.cell(row=row_num, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    row_num += 1

                    # Data rows
                    for row in section['rows']:
                        for col_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=row_num, column=col_idx, value=value)
                            cell.font = cell_font
                            cell.border = thin_border

                            # Format numbers
                            if isinstance(value, (int, float, Decimal)):
                                cell.alignment = Alignment(horizontal='right')
                                if isinstance(value, float):
                                    cell.number_format = '#,##0.00'
                        row_num += 1

                    row_num += 1  # Space between sections

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_csv(
        self,
        headers: List[str],
        rows: List[List[Any]],
    ) -> str:
        """Generate CSV content."""
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([self._format_value(cell) for cell in row])

        return output.getvalue()


# Report type definitions
REPORT_TYPES = {
    # Work Order Reports
    'wo_summary': {
        'name': 'Work Order Summary',
        'description': 'Overview of work orders by status, type, and priority',
        'category': 'Work Orders',
    },
    'wo_completion': {
        'name': 'Work Order Completion Report',
        'description': 'Detailed list of completed work orders with times and costs',
        'category': 'Work Orders',
    },
    'wo_backlog': {
        'name': 'Work Order Backlog Report',
        'description': 'Open work orders aged by creation date',
        'category': 'Work Orders',
    },
    'wo_cost_analysis': {
        'name': 'Work Order Cost Analysis',
        'description': 'Labor and material costs by work order',
        'category': 'Work Orders',
    },
    'wo_technician': {
        'name': 'Technician Performance Report',
        'description': 'Work orders and hours by technician',
        'category': 'Work Orders',
    },

    # Asset Reports
    'asset_summary': {
        'name': 'Asset Summary Report',
        'description': 'Overview of assets by status and criticality',
        'category': 'Assets',
    },
    'asset_reliability': {
        'name': 'Asset Reliability Report (MTBF/MTTR)',
        'description': 'Mean time between failures and repair times by asset',
        'category': 'Assets',
    },
    'asset_downtime': {
        'name': 'Asset Downtime Report',
        'description': 'Total downtime hours by asset',
        'category': 'Assets',
    },
    'asset_cost': {
        'name': 'Asset Cost Report',
        'description': 'Total maintenance costs by asset (Bad Actors)',
        'category': 'Assets',
    },

    # PM Reports
    'pm_compliance': {
        'name': 'PM Compliance Report',
        'description': 'Preventive maintenance compliance rates',
        'category': 'Preventive Maintenance',
    },
    'pm_schedule': {
        'name': 'PM Schedule Report',
        'description': 'Upcoming preventive maintenance schedule',
        'category': 'Preventive Maintenance',
    },

    # Labor Reports
    'labor_summary': {
        'name': 'Labor Summary Report',
        'description': 'Hours and costs by technician',
        'category': 'Labor',
    },
    'labor_by_craft': {
        'name': 'Labor by Craft Report',
        'description': 'Hours and costs by craft/trade',
        'category': 'Labor',
    },
    'overtime_report': {
        'name': 'Overtime Report',
        'description': 'Overtime hours by technician',
        'category': 'Labor',
    },

    # Inventory Reports
    'inventory_value': {
        'name': 'Inventory Value Report',
        'description': 'Current inventory value by storeroom',
        'category': 'Inventory',
    },
    'inventory_usage': {
        'name': 'Parts Usage Report',
        'description': 'Most and least used parts',
        'category': 'Inventory',
    },
    'inventory_reorder': {
        'name': 'Reorder Report',
        'description': 'Parts below reorder point',
        'category': 'Inventory',
    },
}
